import io
import csv
from datetime import datetime
from collections import OrderedDict
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, KeepTogether, Image, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from openpyxl import Workbook
from django.contrib.staticfiles import finders
from .models import Posto, Pessoal # Importe os modelos necessários
from django.utils.html import strip_tags
from django.contrib.auth import get_user_model
import logging

logger = logging.getLogger(__name__)

# Log a message to confirm module starts loading
logger.info("Loading export_utils.py module...")


# --- SEU MAPA DE POSTO/SEÇÃO (MANTIDO) ---
POSTO_SECAO_MAP = {
    "EM": "EM",
    "703150000 - CMT": "CMT",
    "703159000 - SUB CMT": "SUB CMT",
    "703159100 - SEC ADM": "SEC ADM",
    "703159110 - B/1 E B/5": "B/1 E B/5",
    "703159110-1 - B/5": "B/5",
    "703159120 - AA": "AA",
    "703159130 - B/3 E MOTOMEC": "B/3 E MOTOMEC",
    "703159130-1 - MOTOMEC": "MOTOMEC",
    "703159131 - COBOM": "COBOM",
    "703159140 - B/4": "B/4",
    "703159150 - ST UGE": "ST UGE",
    "703159160 - ST PJMD": "ST PJMD",
    "703159200 - SEC ATIV TEC": "SEC ATIV TEC",
    "1ºSGB": "1ºSGB",
    "703151000 - CMT 1º SGB": "CMT 1º SGB",
    "703151100 - ADM PB CERRADO": "ADM PB CERRADO",
    "703151101 - EB CERRADO": "EB CERRADO",
    "703151102 - EB ZONA NORTE": "EB ZONA NORTE",
    "703151200 - ADM PB SANTA ROSÁLIA": "ADM PB SANTA ROSÁLIA",
    "703151201 - EB SANTA ROSÁLIA": "EB SANTA ROSÁLIA",
    "703151202 - EB ÉDEM": "EB ÉDEM",
    "703151300 - ADM PB VOTORANTIM": "ADM PB VOTORANTIM",
    "703151301 - EB VOTORANTIM": "EB VOTORANTIM",
    "703151302 - EB PIEDADE": "EB PIEDADE",
    "703151800 - ADM 1º SGB": "ADM 1º SGB",
    "2ºSGB": "2ºSGB",
    "703152000 - CMT 2º SGB": "CMT 2º SGB",
    "703152100 - ADM PB ITU": "ADM PB ITU",
    "703152101 - EB ITU": "EB ITU",
    "703152102 - EB PORTO FELIZ": "EB PORTO FELIZ",
    "703152200 - ADM PB SALTO": "ADM PB SALTO",
    "703152201 - EB SALTO": "EB SALTO",
    "703152300 - ADM PB SÃO ROQUE": "ADM PB SÃO ROQUE",
    "703152301 - EB SÃO ROQUE": "EB SÃO ROQUE",
    "703152302 - EB IBIÚNA": "EB IBIÚNA",
    "703152800 - ADM 2º SGB": "ADM 2º SGB",
    "703152900 - NUCL ATIV TEC 2º SGB": "NUCL ATIV TEC 2º SGB",
    "3ºSGB": "3ºSGB",
    "703153000 - CMT 3º SGB": "CMT 3º SGB",
    "703153100 - ADM PB ITAPEVA": "ADM PB ITAPEVA",
    "703153101 - EB ITAPEVA": "EB ITAPEVA",
    "703153102 - EB APIAÍ": "EB APIAÍ",
    "703153103 - EB ITARARÉ": "EB ITARARÉ",
    "703153104 - EB CAPÃO BONITO": "EB CAPÃO BONITO",
    "703153800 - ADM 3º SGB": "ADM 3º SGB",
    "703153900 - NUCL ATIV TEC 3º SGB": "NUCL ATIV TEC 3º SGB",
    "4ºSGB": "4ºSGB",
    "703154000 - CMT 4º SGB": "CMT 4º SGB",
    "703154100 - ADM PB ITAPETININGA": "ADM PB ITAPETININGA",
    "703154101 - EB ITAPETININGA": "EB ITAPETININGA",
    "703154102 - EB BOITUVA": "EB BOITUVA",
    "703154103 - EB ANGATUBA": "EB ANGATUBA",
    "703154200 - ADM PB TATUÍ": "ADM PB TATUÍ",
    "703154201 - EB TATUÍ": "EB TATUÍ",
    "703154202 - EB TIETÊ": "EB TIETÊ",
    "703154203 - EB LARANJAL PAULISTA": "EB LARANJAL PAULISTA",
    "703154800 - ADM 4º SGB": "ADM 4º SGB",
    "703154900 - NUCL ATIV TEC 4º SGB": "NUCL ATIV TEC 4º SGB",
    "5ºSGB": "5ºSGB",
    "703155000 - CMT 5º SGB": "CMT 5º SGB",
    "703155100 - ADM PB BOTUCATU": "ADM PB BOTUCATU",
    "703155101 - EB BOTUCATU": "EB BOTUCATU",
    "703155102 - EB ITATINGA": "EB ITATINGA",
    "703155200 - ADM PB AVARÉ": "ADM PB AVARÉ",
    "703155201 - EB AVARÉ": "EB AVARÉ",
    "703155202 - EB PIRAJU": "EB PIRAJU",
    "703155203 - EB ITAÍ": "EB ITAÍ",
    "703155800 - ADM 5º SGB": "ADM 5º SGB",
    "703155900 - NUCL ATIV TEC 5º SGB": "NUCL ATIV TEC 5º SGB"
}

# --- FUNÇÃO DISPATCHER (MANTIDA) ---
def export_rpt_data(request, export_format='xlsx', **filters):
    logger.debug(f"Filtros: {filters}")

    # Moved import inside the function to avoid module-level issues if Cadastro_rpt is problematic
    Cadastro_rpt = None
    try:
        from backend.rpt.models import Cadastro_rpt
    except ImportError:
        logger.warning("Modelo Cadastro_rpt não encontrado. A exportação RPT pode não funcionar.")
        # Cadastro_rpt remains None

    queryset = []
    if Cadastro_rpt:
        queryset = Cadastro_rpt.objects.all()
        # Garantir que o filtro de status seja sempre "Aguardando"
        queryset = queryset.filter(status='Aguardando')

        if filters.get('posto_secao_destino'):
            queryset = queryset.filter(posto_secao_destino=filters['posto_secao_destino'])

        logger.debug(f"Queryset: {queryset.query}")
        logger.debug(f"Número de registros encontrados: {queryset.count()}")
    else:
        logger.info("Não foi possível carregar Cadastro_rpt, retornando queryset vazio para RPT.")


    if export_format == 'pdf':
        buffer_bytes, filename = export_to_pdf_rpt(request, queryset)
        content_type = 'application/pdf'

    elif export_format == 'xlsx':
        buffer_bytes, filename = export_to_excel_rpt(request, queryset)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    elif export_format == 'csv':
        buffer_bytes, filename = export_to_csv_rpt(request, queryset)
        content_type = 'text/csv; charset=utf-8-sig'

    else:
        raise ValueError("Formato não suportado")

    response = HttpResponse(buffer_bytes, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO PDF COM MARCA D'ÁGUA (RPT)
# ==============================================
def export_to_pdf_rpt(request, data):
    """Gera o relatório RPT em PDF com marca d'água."""
    buffer = io.BytesIO()
    user = request.user # Usuário logado que está emitindo

    # Configura estilos básicos
    styles = getSampleStyleSheet()
    # Renomeando estilos para evitar conflitos
    styles.add(ParagraphStyle(
        name='RPTHeader', # Renomeado
        fontSize=11,
        leading=13,
        alignment=1, # Centralizado
        spaceAfter=4*mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#002060')
    ))
    styles.add(ParagraphStyle(
        name='RPTFooterInfo', # Renomeado
        fontSize=8,
        alignment=1, # Centralizado
        textColor=colors.grey
    ))
    styles.add(ParagraphStyle(
        name='RPTSGB', # Renomeado
        fontSize=11,
        leading=13,
        spaceBefore=6*mm,
        textColor=colors.HexColor('#004080'),
        fontName='Helvetica-Bold'
    ))
    styles.add(ParagraphStyle(
        name='RPTPosto', # Renomeado
        fontSize=10,
        leading=12,
        spaceBefore=1*mm,
        textColor=colors.HexColor('#404040')
    ))
    styles.add(ParagraphStyle(
            name='RPTTotalPosto', # Renomeado
            fontSize=8,
            alignment=2, # Alinhado à direita
            spaceBefore=1*mm,
            textColor=colors.dimgray
     ))

    # Configura o documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10*mm,
        rightMargin=10*mm,
        topMargin=15*mm,
        bottomMargin=20*mm
    )

    elements = []

    # --- Cabeçalho com Logos ---
    try:
        def get_image_path(file):
            path = finders.find(f'img/{file}')
            if not path:
                raise FileNotFoundError(f"Arquivo estático img/{file} não encontrado!")
            return path

        logo_path = get_image_path('logo.png')
        brasao_path = get_image_path('brasao.png')

        logo_esquerda = Image(logo_path, width=12*mm, height=20*mm)
        logo_direita = Image(brasao_path, width=20*mm, height=20*mm)

        header_text = Paragraph(
            "Relação de Prioridade de Transferência<br/>15º Grupamento de Bombeiros Militar",
            styles['RPTHeader'] # Usando o estilo renomeado
        )

        header_table = Table([
            [logo_esquerda, header_text, logo_direita]
        ], colWidths=[doc.width*0.15, doc.width*0.7, doc.width*0.15])

        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'CENTER'),
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 8*mm))
    except Exception as e:
        elements.append(Paragraph(f"Erro ao gerar cabeçalho: {str(e)}", styles['Normal']))

    # --- Processamento e Agrupamento dos Dados ---
    organized_data = OrderedDict()
    for item in data:
        sgb_key = item.sgb_destino
        sgb_display = POSTO_SECAO_MAP.get(item.sgb_destino, item.sgb_destino)
        posto_key = item.posto_secao_destino
        posto_display = POSTO_SECAO_MAP.get(item.posto_secao_destino, item.posto_secao_destino)

        if sgb_display not in organized_data:
            organized_data[sgb_display] = OrderedDict()
        if posto_display not in organized_data[sgb_display]:
            organized_data[sgb_display][posto_display] = []
        organized_data[sgb_display][posto_display].append(item)

    total_geral = 0

    table_header = [
        'Posto/Grad', 'RE', 'Nome Guerra',
        'SGB Origem', 'Posto Origem',
        'Dt. Pedido',
        'Dias'
    ]

    table_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E0E0E0')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 3*mm),

        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
        ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 7),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),

        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('LEFTPADDING', (0,0), (-1,-1), 1.5*mm),
        ('RIGHTPADDING', (0,0), (-1,-1), 1.5*mm),
        ('TOPPADDING', (0,0), (-1,-1), 1*mm),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1*mm),
    ])

    for sgb_display, postos in organized_data.items():
        sgb_elements = []
        sgb_elements.append(Paragraph(sgb_display, styles['RPTSGB'])) # Usando o estilo renomeado

        for posto_display, items in postos.items():
            posto_elements = []
            posto_elements.append(Paragraph(posto_display, styles['RPTPosto'])) # Usando o estilo renomeado
            posto_elements.append(Spacer(1, 1.5*mm))

            table_data = [table_header]
            count_posto = 0
            for item in items:
                # Acessa dados relacionados com segurança (verifica existência)
                # Usa strip_tags para remover HTML de `grad` se houver
                posto_grad = strip_tags(item.cadastro.promocoes.last().grad) if hasattr(item.cadastro, 'promocoes') and item.cadastro.promocoes.exists() else '-'

                # Acessa detalhes da situação com segurança
                last_situacao = item.cadastro.detalhes_situacao.last() if hasattr(item.cadastro, 'detalhes_situacao') and item.cadastro.detalhes_situacao.exists() else None

                sgb_origem = POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb) if last_situacao else '-'
                posto_origem = POSTO_SECAO_MAP.get(last_situacao.posto_secao, last_situacao.posto_secao) if last_situacao else '-'

                table_data.append([
                    posto_grad,
                    f'{item.cadastro.re}-{item.cadastro.dig}',
                    item.cadastro.nome_de_guerra,
                    sgb_origem,
                    posto_origem,
                    item.data_pedido.strftime('%d/%m/%Y'),
                    str((datetime.now().date() - item.data_pedido).days)
                ])
                count_posto += 1

            table = Table(table_data,
                          colWidths=[25*mm, 15*mm, 40*mm, 20*mm, 55*mm, 20*mm, 10*mm],
                          style=table_style)

            posto_elements.append(table)
            posto_elements.append(Paragraph(f"Total {posto_display}: {count_posto}", styles['RPTTotalPosto'])) # Usando o estilo renomeado
            posto_elements.append(Spacer(1, 4*mm))

            sgb_elements.extend(posto_elements)
            total_geral += count_posto

        elements.append(KeepTogether(sgb_elements))
        elements.append(Spacer(1, 6*mm))

    # Log a message to confirm function is being defined
    logger.info("Defining export_efetivo_pdf_report function...")

# ==============================================
# NOVA FUNÇÃO PARA EXPORTAÇÃO PDF DE RELATÓRIO DE EFETIVO
# ==============================================
def export_efetivo_pdf_report(request, report_data, sgb_filter=None, posto_secao_filter=None):
    """
    Gera um relatório PDF do efetivo, agrupado por SGB e Posto/Seção,
    no formato especificado pelo usuário.
    """
    buffer = io.BytesIO()
    user = request.user

    # Definir os grupos para contagem do efetivo existente
    GRUPOS = {
        'Tc': 'Ten Cel',
        'Maj': 'Maj',
        'Cap': 'Cap',
        'Ten': 'Ten QO',
        'Ten QAOPM': 'Ten QA',
        'St/Sgt': 'St/Sgt',
        'Cb/Sd': 'Cb/Sd'
    }

    # --- Estilos ReportLab ---
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='CustomTitle',
        fontSize=14,
        leading=18,
        alignment=1, # CENTER
        spaceAfter=10*mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#002060')
    ))
    styles.add(ParagraphStyle(
        name='CustomSubtitle',
        fontSize=10,
        leading=12,
        alignment=1, # CENTER
        spaceAfter=5*mm,
        textColor=colors.gray
    ))
    styles.add(ParagraphStyle(
        name='CustomSectionHeaderSGB',
        fontSize=12,
        leading=14,
        spaceBefore=8*mm,
        spaceAfter=3*mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#004080')
    ))
    styles.add(ParagraphStyle(
        name='CustomSectionHeaderPostoSecao',
        fontSize=10,
        leading=12,
        spaceBefore=4*mm,
        spaceAfter=2*mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#404040')
    ))
    styles.add(ParagraphStyle(
        name='CustomTableData',
        fontSize=8,
        alignment=1, # CENTER
        textColor=colors.black
    ))
    styles.add(ParagraphStyle(
        name='CustomTableDataBold',
        fontSize=8,
        alignment=1, # CENTER
        fontName='Helvetica-Bold',
        textColor=colors.black
    ))
    styles.add(ParagraphStyle(
        name='CustomTableDataGreen',
        fontSize=8,
        alignment=1, # CENTER
        textColor=colors.green
    ))
    styles.add(ParagraphStyle(
        name='CustomTableDataRed',
        fontSize=8,
        alignment=1, # CENTER
        textColor=colors.red
    ))
    styles.add(ParagraphStyle(
        name='CustomFooterInfo',
        fontSize=8,
        alignment=1, # CENTER
        textColor=colors.grey
    ))

    # --- Document Template ---
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    elements = []

    # --- Cabeçalho Global do Relatório ---
    elements.append(Paragraph("Relatório de Efetivo por Posto/Seção", styles['CustomTitle']))
    elements.append(Paragraph(f"Data de Emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['CustomSubtitle']))
    
    # Adicionar filtros ao subtítulo se aplicáveis
    filter_info = []
    if sgb_filter:
        filter_info.append(f"SGB: {POSTO_SECAO_MAP.get(sgb_filter, sgb_filter)}")
    if posto_secao_filter:
        filter_info.append(f"Posto/Seção: {POSTO_SECAO_MAP.get(posto_secao_filter, posto_secao_filter)}")
    
    if filter_info:
        elements.append(Paragraph("Filtros: " + ", ".join(filter_info), styles['CustomSubtitle']))
    elements.append(Spacer(1, 10*mm))


    # --- Agrupamento dos Dados para o Relatório ---
    grouped_data = OrderedDict()
    for item in report_data:
        posto_obj = item['posto_obj']
        
        sgb_display = posto_obj.get_sgb_display()
        posto_secao_display = posto_obj.get_posto_secao_display()

        if sgb_display not in grouped_data:
            grouped_data[sgb_display] = OrderedDict()
        if posto_secao_display not in grouped_data[sgb_display]:
            grouped_data[sgb_display][posto_secao_display] = item

    # --- Geração do Conteúdo do Relatório ---
    total_geral_efetivo_existente = 0
    total_geral_pessoal_planejado = 0

    for sgb_display, posto_secoes in grouped_data.items():
        elements.append(Paragraph(f"SGB: {sgb_display}", styles['CustomSectionHeaderSGB']))
        elements.append(Spacer(1, 2*mm))

        for posto_secao_display, item in posto_secoes.items():
            posto_obj = item['posto_obj']
            
            # Recalcular efetivo_existente_grupos e total_efetivo_existente aqui
            # Importar modelos de efetivo dentro da função para evitar ImportError no módulo
            DetalhesSituacao = None
            Promocao = None
            try:
                from backend.efetivo.models import DetalhesSituacao, Promocao
            except ImportError:
                logger.warning("Modelos DetalhesSituacao ou Promocao não encontrados. Efetivo Existente pode não ser calculado.")

            efetivo_existente_grupos = {grupo: 0 for grupo in GRUPOS.keys()}
            total_efetivo_existente = 0

            if DetalhesSituacao and Promocao:
                efetivos_existentes = DetalhesSituacao.objects.filter(
                    situacao='Efetivo',
                    posto_secao=posto_obj.posto_secao # Filtrar pelo posto_secao do posto atual
                ).select_related('cadastro').prefetch_related('cadastro__promocoes') # Prefetch para evitar N+1 queries

                for ef_existente in efetivos_existentes:
                    ultima_promocao = ef_existente.cadastro.promocoes.order_by('-ultima_promocao').first()
                    if ultima_promocao:
                        # Mapear o 'grupo' da promoção para as chaves do dicionário GRUPOS
                        grupo_key = ultima_promocao.grupo.strip()
                        # Ajuste para 'Ten QAOPM' se necessário, ou garanta que o grupo seja mapeado corretamente
                        if grupo_key == 'Ten QAOPM':
                            grupo_key = 'Ten QAOPM' # Manter como está, ou ajustar para 'Ten_QAOPM' se a chave for essa
                        elif grupo_key == 'St/Sgt':
                            grupo_key = 'St/Sgt'
                        elif grupo_key == 'Cb/Sd':
                            grupo_key = 'Cb/Sd'
                        elif grupo_key == 'Ten': # Para Ten QO
                            grupo_key = 'Ten'
                        
                        if grupo_key in efetivo_existente_grupos:
                            efetivo_existente_grupos[grupo_key] += 1
                total_efetivo_existente = sum(efetivo_existente_grupos.values())
            else:
                logger.info(f"Não foi possível calcular efetivo existente para {posto_obj.posto_atendimento} devido a modelos ausentes.")

            pessoal_planejado = item['pessoal_planejado']

            elements.append(Paragraph(f"Posto/Seção: {posto_secao_display} - {posto_obj.posto_atendimento}", styles['CustomSectionHeaderPostoSecao']))
            elements.append(Spacer(1, 2*mm))

            # Tabela de Efetivo Fixado (QPO)
            elements.append(Paragraph("Efetivo Fixado (QPO)", styles['Normal'])) # Título alterado
            table_fixado_header = [
                Paragraph("Cel", styles['CustomTableDataBold']), # Adicionado Cel
                Paragraph("Ten Cel", styles['CustomTableDataBold']),
                Paragraph("Maj", styles['CustomTableDataBold']),
                Paragraph("Cap", styles['CustomTableDataBold']),
                Paragraph("Ten QO", styles['CustomTableDataBold']),
                Paragraph("Ten QA", styles['CustomTableDataBold']),
                Paragraph("Asp", styles['CustomTableDataBold']), # Adicionado Asp
                Paragraph("St/Sgt", styles['CustomTableDataBold']),
                Paragraph("Cb/Sd", styles['CustomTableDataBold']),
                Paragraph("Total", styles['CustomTableDataBold'])
            ]
            table_fixado_data = [
                table_fixado_header,
                [
                    Paragraph(str(pessoal_planejado.get('cel', 0)), styles['CustomTableData']), # Dados de Cel
                    Paragraph(str(pessoal_planejado.get('ten_cel', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('maj', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('cap', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('tenqo', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('tenqa', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('asp', 0)), styles['CustomTableData']), # Dados de Asp
                    Paragraph(str(pessoal_planejado.get('st_sgt', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('cb_sd', 0)), styles['CustomTableData']),
                    Paragraph(str(pessoal_planejado.get('total_planejado', 0)), styles['CustomTableDataBold'])
                ]
            ]
            # Ajuste de colWidths para a tabela "Efetivo Fixado" (10 colunas)
            # A4 width - 2*margin = 210mm - 30mm = 180mm
            # 180mm / 10 colunas = 18mm por coluna
            col_widths_fixado = [18*mm] * 10 
            table_fixado_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E0E0E0')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('LEFTPADDING', (0,0), (-1,-1), 1*mm), # Reduzir padding
                ('RIGHTPADDING', (0,0), (-1,-1), 1*mm), # Reduzir padding
                ('TOPPADDING', (0,0), (-1,-1), 1*mm),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1*mm),
            ])
            table_fixado = Table(table_fixado_data, colWidths=col_widths_fixado, style=table_fixado_style)
            elements.append(table_fixado)
            elements.append(Spacer(1, 4*mm))

            # Tabela de Efetivo Existente
            elements.append(Paragraph("Efetivo Existente", styles['Normal'])) # Título alterado
            table_existente_header = [ # Atualizado para refletir os grupos que você tem
                Paragraph("Ten Cel", styles['CustomTableDataBold']),
                Paragraph("Maj", styles['CustomTableDataBold']),
                Paragraph("Cap", styles['CustomTableDataBold']),
                Paragraph("Ten QO", styles['CustomTableDataBold']),
                Paragraph("Ten QA", styles['CustomTableDataBold']),
                Paragraph("St/Sgt", styles['CustomTableDataBold']),
                Paragraph("Cb/Sd", styles['CustomTableDataBold']),
                Paragraph("Total", styles['CustomTableDataBold'])
            ]
            table_existente_data = [
                table_existente_header,
                [
                    Paragraph(str(efetivo_existente_grupos.get('Tc', 0)), styles['CustomTableData']),
                    Paragraph(str(efetivo_existente_grupos.get('Maj', 0)), styles['CustomTableData']),
                    Paragraph(str(efetivo_existente_grupos.get('Cap', 0)), styles['CustomTableData']),
                    Paragraph(str(efetivo_existente_grupos.get('Ten', 0)), styles['CustomTableData']),
                    Paragraph(str(efetivo_existente_grupos.get('Ten QAOPM', 0)), styles['CustomTableData']), # Usar a chave correta
                    Paragraph(str(efetivo_existente_grupos.get('St/Sgt', 0)), styles['CustomTableData']), # Usar a chave correta
                    Paragraph(str(efetivo_existente_grupos.get('Cb/Sd', 0)), styles['CustomTableData']), # Usar a chave correta
                    Paragraph(str(total_efetivo_existente), styles['CustomTableDataBold'])
                ]
            ]
            # Ajuste de colWidths para a tabela "Efetivo Existente" (8 colunas)
            # 180mm / 8 colunas = 22.5mm por coluna
            col_widths_existente = [22.5*mm] * 8
            table_existente_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E0E0E0')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('LEFTPADDING', (0,0), (-1,-1), 2*mm),
                ('RIGHTPADDING', (0,0), (-1,-1), 2*mm),
                ('TOPPADDING', (0,0), (-1,-1), 1*mm),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1*mm),
            ])
            table_existente = Table(table_existente_data, colWidths=col_widths_existente, style=table_existente_style)
            elements.append(table_existente)
            elements.append(Spacer(1, 4*mm))

            # Tabela de Porcentagem de Diferença (mantida)
            elements.append(Paragraph("Porcentagem de Diferença", styles['Normal']))
            table_diff_header = [
                Paragraph("Ten Cel", styles['CustomTableDataBold']),
                Paragraph("Maj", styles['CustomTableDataBold']),
                Paragraph("Cap", styles['CustomTableDataBold']),
                Paragraph("Ten QO", styles['CustomTableDataBold']),
                Paragraph("Ten QA", styles['CustomTableDataBold']),
                Paragraph("St/Sgt", styles['CustomTableDataBold']),
                Paragraph("Cb/Sd", styles['CustomTableDataBold']),
                Paragraph("Total", styles['CustomTableDataBold'])
            ]
            
            diff_row = []
            # Calcula a diferença e a porcentagem
            # Mapeamento para as chaves corretas do efetivo existente e planejado
            mapping_keys = [
                ('Tc', 'ten_cel'), ('Maj', 'maj'), ('Cap', 'cap'), ('Ten', 'tenqo'),
                ('Ten QAOPM', 'tenqa'), ('St/Sgt', 'st_sgt'), ('Cb/Sd', 'cb_sd')
            ]

            for key_existente, key_planejado in mapping_keys:
                existente_val = efetivo_existente_grupos.get(key_existente, 0)
                planejado_val = pessoal_planejado.get(key_planejado, 0)
                
                diff = existente_val - planejado_val
                
                if planejado_val == 0:
                    percentage = 100 if existente_val > 0 else 0
                else:
                    percentage = (diff / planejado_val) * 100
                
                style_diff = styles['CustomTableData']
                if diff > 0:
                    style_diff = styles['CustomTableDataGreen']
                elif diff < 0:
                    style_diff = styles['CustomTableDataRed']

                diff_row.append(Paragraph(f"{diff} ({percentage:.2f}%)", style_diff))

            # Total da Porcentagem de Diferença
            total_diff = total_efetivo_existente - pessoal_planejado.get('total_planejado', 0)
            if pessoal_planejado.get('total_planejado', 0) == 0:
                total_percentage = 100 if total_efetivo_existente > 0 else 0
            else:
                total_percentage = (total_diff / pessoal_planejado.get('total_planejado', 0)) * 100

            style_total_diff = styles['CustomTableDataBold']
            if total_diff > 0:
                style_total_diff = styles['CustomTableDataGreen']
            elif total_diff < 0:
                style_total_diff = styles['CustomTableDataRed']

            diff_row.append(Paragraph(f"{total_diff} ({total_percentage:.2f}%)", style_total_diff))

            table_diff_data = [
                table_diff_header,
                diff_row
            ]
            table_diff_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E0E0E0')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('LEFTPADDING', (0,0), (-1,-1), 2*mm),
                ('RIGHTPADDING', (0,0), (-1,-1), 2*mm),
                ('TOPPADDING', (0,0), (-1,-1), 1*mm),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1*mm),
            ])
            table_diff = Table(table_diff_data, colWidths=col_widths_existente, style=table_diff_style) # Usar col_widths_existente
            elements.append(table_diff)
            elements.append(Spacer(1, 8*mm))

            if posto_secao_display != list(posto_secoes.keys())[-1] or sgb_display != list(grouped_data.keys())[-1]:
                 elements.append(PageBreak())


    # --- Funções para Rodapé e Marca D'água (Ajustadas para este relatório) ---
    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)

        User = get_user_model()
        user_obj = User.objects.get(pk=user.pk)
        
        posto_grad = getattr(getattr(user_obj, 'profile', None), 'posto_grad', '')
        re = getattr(getattr(user_obj, 'profile', None), 're', '')
        dig = getattr(user_obj, 'dig', '')
        last_name = getattr(user_obj, 'last_name', '')
        
        user_info = f"{posto_grad} {re}-{dig} {last_name}".strip()
        if not user_info and hasattr(user_obj, 'username'):
            user_info = user_obj.username
        elif not user_info and hasattr(user_obj, 'email'):
            user_info = user_obj.email

        footer_text = (
            f"Emitido por: {user_info} | "
            f"{datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"Página {doc.page}"
        )
        canvas.drawCentredString(A4[0]/2, 10*mm, footer_text)
        canvas.restoreState()

    def add_watermark(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 40)
        canvas.setFillGray(0.9, 0.15)

        User = get_user_model()
        user_obj = User.objects.get(pk=user.pk)
        cpf = getattr(getattr(user_obj, 'profile', None), 'cpf', '')
        text = cpf if cpf else "CONFIDENCIAL"

        angle = 45
        for x in range(-int(A4[1]), int(A4[0] * 1.5), 250):
            for y in range(-int(A4[0]), int(A4[1] * 1.5), 150):
                canvas.saveState()
                canvas.translate(x, y)
                canvas.rotate(angle)
                canvas.drawCentredString(0, 0, text)
                canvas.restoreState()
        canvas.restoreState()

    def on_page(canvas, doc):
        footer(canvas, doc)
        add_watermark(canvas, doc)

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    pdf_buffer = buffer.getvalue()
    buffer.close()
    return pdf_buffer, "relatorio_efetivo.pdf" # Nome do arquivo para o relatório de efetivo

# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO EXCEL (Mantida)
# ==============================================
def export_to_excel_rpt(request, data):
    buffer = io.BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "RPT - Prioridade Transferência"

    ws.append(["RELATÓRIO DE PRIORIDADE DE TRANSFERÊNCIA - 15º GBM"])
    ws.append([f"Data de emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    emitter = request.user.get_full_name() or request.user.username
    ws.append([f"Emitido por: {emitter}"])
    ws.append([])

    headers = [
        'Posto/Grad', 'RE', 'Nome Guerra',
        'SGB Origem', 'Posto Origem',
        'Data Pedido', 'Dias Espera'
    ]
    ws.append(headers)

    for item in data:
        posto_grad = strip_tags(item.cadastro.promocoes.last().grad) if hasattr(item.cadastro, 'promocoes') and item.cadastro.promocoes.exists() else '-'
        last_situacao = item.cadastro.detalhes_situacao.last() if hasattr(item.cadastro, 'detalhes_situacao') and item.cadastro.detalhes_situacao.exists() else None

        sgb_origem = POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb) if last_situacao else '-'
        posto_origem = POSTO_SECAO_MAP.get(last_situacao.posto_secao, last_situacao.posto_secao) if last_situacao else '-'

        ws.append([
            posto_grad,
            f'{item.cadastro.re}-{item.cadastro.dig}',
            item.cadastro.nome_de_guerra,
            sgb_origem,
            posto_origem,
            item.data_pedido,
            (datetime.now().date() - item.data_pedido).days
        ])

    column_widths = {'A': 15, 'B': 10, 'C': 30, 'D': 15, 'E': 40, 'F': 15, 'G': 10}
    for col_letter, width in column_widths.items():
         if col_letter in ws.column_dimensions:
              ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=6, max_col=7, max_row=ws.max_row):
        if row[5].value:
             row[5].number_format = 'dd/mm/yyyy'
        if row[6].value is not None:
             row[6].number_format = '0'


    wb.save(buffer)
    buffer_bytes = buffer.getvalue()
    buffer.close()
    
    return buffer_bytes, "relatorio_prioridade_transferencia.xlsx"


# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO CSV (Mantida)
# ==============================================
def export_to_csv_rpt(request, data):
    buffer = io.BytesIO()
    buffer.write(b'\xEF\xBB\xBF')

    wrapper = io.TextIOWrapper(
        buffer,
        encoding='utf-8-sig',
        newline='',
        write_through=True
    )

    writer = csv.writer(wrapper, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    try:
        writer.writerow(["RELATÓRIO DE PRIORIDADE DE TRANSFERÊNCIA - 15º GBM"])
        writer.writerow([f"Data de emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        emitter = request.user.get_full_name() or request.user.username
        writer.writerow([f"Emitido por: {emitter}"])
        writer.writerow([])

        headers = [
            'Posto/Grad', 'RE', 'Nome Guerra',
            'SGB Origem', 'Posto Origem',
            'Data Pedido', 'Dias Espera'
        ]
        writer.writerow(headers)

        for item in data:
            posto_grad = strip_tags(item.cadastro.promocoes.last().grad) if hasattr(item.cadastro, 'promocoes') and item.cadastro.promocoes.exists() else '-'
            last_situacao = item.cadastro.detalhes_situacao.last() if hasattr(item.cadastro, 'detalhes_situacao') and item.cadastro.detalhes_situacao.exists() else None

            sgb_origem = POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb) if last_situacao else '-'
            posto_origem = POSTO_SECAO_MAP.get(last_situacao.posto_secao, last_situacao.posto_secao) if last_situacao else '-'

            writer.writerow([
                posto_grad,
                f'{item.cadastro.re}-{item.cadastro.dig}',
                item.cadastro.nome_de_guerra,
                sgb_origem,
                posto_origem,
                item.data_pedido.strftime('%d/%m/%Y'),
                (datetime.now().date() - item.data_pedido).days
            ])

    finally:
        wrapper.detach()

    buffer_bytes = buffer.getvalue()
    buffer.close()
    
    return buffer_bytes, "relatorio_prioridade_transferencia.csv"
