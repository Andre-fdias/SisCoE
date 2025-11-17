import io
import csv
from datetime import datetime
from collections import OrderedDict
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    KeepTogether,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from openpyxl import Workbook
from django.contrib.staticfiles import finders
from .models import Cadastro_rpt  # Assuming Cadastro_rpt is in the same app's models

# Import DetalhesSituacao if needed directly, but likely accessed via Cadastro_rpt.cadastro
# from backend.efetivo.models import DetalhesSituacao
from django.utils.html import strip_tags

# Import needed for watermark user info
import logging

logger = logging.getLogger(__name__)


# --- SEU MAPA DE POSTO/SEÇÃO ---
POSTO_SECAO_MAP = {
    # ... (seu mapa completo aqui) ...
    "EM": "EM",
    "703150000 - CMT": "CMT",
    "703159000 - SUB CMT": "SUB CMT",
    "703159100 - SEC ADM": "SEC ADM",
    "703159110 - B/1 E B/5": "B/1 E B/5",
    "703159110-1 - B/5": "B/5",
    "703159120 - AA": "AA",
    "703159130 - B/3 E MOTOMEC": "B/3 E MOTOMEC",
    "703159130-1 - MOTOMEC": "MOTOMEC",
    "703159131 - COBOM": "COBOM",
    "703159140 - B/4": "B/4",
    "703159150 - ST UGE": "ST UGE",
    "703159160 - ST PJMD": "ST PJMD",
    "703159200 - SEC ATIV TEC": "SEC ATIV TEC",
    "1ºSGB": "1ºSGB",
    "703151000 - CMT 1º SGB": "CMT 1º SGB",
    "703151100 - ADM PB CERRADO": "ADM PB CERRADO",
    "703151101 - EB CERRADO": "EB CERRADO",
    "703151102 - EB ZONA NORTE": "EB ZONA NORTE",
    "703151200 - ADM PB SANTA ROSÁLIA": "ADM PB SANTA ROSÁLIA",
    "703151201 - EB SANTA ROSÁLIA": "EB SANTA ROSÁLIA",
    "703151202 - EB ÉDEM": "EB ÉDEM",
    "703151300 - ADM PB VOTORANTIM": "ADM PB VOTORANTIM",
    "703151301 - EB VOTORANTIM": "EB VOTORANTIM",
    "703151302 - EB PIEDADE": "EB PIEDADE",
    "703151800 - ADM 1º SGB": "ADM 1º SGB",
    "2ºSGB": "2ºSGB",
    "703152000 - CMT 2º SGB": "CMT 2º SGB",
    "703152100 - ADM PB ITU": "ADM PB ITU",
    "703152101 - EB ITU": "EB ITU",
    "703152102 - EB PORTO FELIZ": "EB PORTO FELIZ",
    "703152200 - ADM PB SALTO": "ADM PB SALTO",
    "703152201 - EB SALTO": "EB SALTO",
    "703152300 - ADM PB SÃO ROQUE": "ADM PB SÃO ROQUE",
    "703152301 - EB SÃO ROQUE": "EB SÃO ROQUE",
    "703152302 - EB IBIÚNA": "EB IBIÚNA",
    "703152800 - ADM 2º SGB": "ADM 2º SGB",
    "703152900 - NUCL ATIV TEC 2º SGB": "NUCL ATIV TEC 2º SGB",
    "3ºSGB": "3ºSGB",
    "703153000 - CMT 3º SGB": "CMT 3º SGB",
    "703153100 - ADM PB ITAPEVA": "ADM PB ITAPEVA",
    "703153101 - EB ITAPEVA": "EB ITAPEVA",
    "703153102 - EB APIAÍ": "EB APIAÍ",
    "703153103 - EB ITARARÉ": "EB ITARARÉ",
    "703153104 - EB CAPÃO BONITO": "EB CAPÃO BONITO",
    "703153800 - ADM 3º SGB": "ADM 3º SGB",
    "703153900 - NUCL ATIV TEC 3º SGB": "NUCL ATIV TEC 3º SGB",
    "4ºSGB": "4ºSGB",
    "703154000 - CMT 4º SGB": "CMT 4º SGB",
    "703154100 - ADM PB ITAPETININGA": "ADM PB ITAPETININGA",
    "703154101 - EB ITAPETININGA": "EB ITAPETININGA",
    "703154102 - EB BOITUVA": "EB BOITUVA",
    "703154103 - EB ANGATUBA": "EB ANGATUBA",
    "703154200 - ADM PB TATUÍ": "ADM PB TATUÍ",
    "703154201 - EB TATUÍ": "EB TATUÍ",
    "703154202 - EB TIETÊ": "EB TIETÊ",
    "703154203 - EB LARANJAL PAULISTA": "EB LARANJAL PAULISTA",
    "703154800 - ADM 4º SGB": "ADM 4º SGB",
    "703154900 - NUCL ATIV TEC 4º SGB": "NUCL ATIV TEC 4º SGB",
    "5ºSGB": "5ºSGB",
    "703155000 - CMT 5º SGB": "CMT 5º SGB",
    "703155100 - ADM PB BOTUCATU": "ADM PB BOTUCATU",
    "703155101 - EB BOTUCATU": "EB BOTUCATU",
    "703155102 - EB ITATINGA": "EB ITATINGA",
    "703155200 - ADM PB AVARÉ": "ADM PB AVARÉ",
    "703155201 - EB AVARÉ": "EB AVARÉ",
    "703155202 - EB PIRAJU": "EB PIRAJU",
    "703155203 - EB ITAÍ": "EB ITAÍ",
    "703155800 - ADM 5º SGB": "ADM 5º SGB",
    "703155900 - NUCL ATIV TEC 5º SGB": "NUCL ATIV TEC 5º SGB",
}
# --- FUNÇÃO DISPATCHER (Aparentemente OK, chama as outras) ---
# Ajuste: Renomeei para refletir que ela recebe request, queryset e format_type da view
# A versão anterior parecia pegar o queryset internamente, o que não é ideal
# A view deve passar o queryset preparado.

# --- FUNÇÃO DISPATCHER CORRIGIDA ---


def export_rpt_data(request, export_format="xlsx", **filters):
    logger.debug(f"Filtros: {filters}")

    queryset = Cadastro_rpt.objects.all()

    # Garantir que o filtro de status seja sempre "Aguardando"
    queryset = queryset.filter(status="Aguardando")

    if filters.get("posto_secao_destino"):
        queryset = queryset.filter(posto_secao_destino=filters["posto_secao_destino"])

    logger.debug(f"Queryset: {queryset.query}")
    logger.debug(f"Número de registros encontrados: {queryset.count()}")

    if export_format == "pdf":
        buffer_bytes, filename = export_to_pdf_rpt(request, queryset)
        content_type = "application/pdf"

    elif export_format == "xlsx":
        buffer_bytes, filename = export_to_excel_rpt(request, queryset)
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif export_format == "csv":
        buffer_bytes, filename = export_to_csv_rpt(request, queryset)
        content_type = "text/csv; charset=utf-8-sig"

    else:
        raise ValueError("Formato não suportado")

    response = HttpResponse(buffer_bytes, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO PDF COM MARCA D'ÁGUA
# ==============================================
def export_to_pdf_rpt(request, data):
    """Gera o relatório RPT em PDF com marca d'água."""
    buffer = io.BytesIO()
    user = request.user  # Usuário logado que está emitindo

    # Configura estilos básicos
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Header",
            fontSize=11,  # Ajustado tamanho
            leading=13,
            alignment=1,  # Centralizado
            spaceAfter=4 * mm,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#002060"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="FooterInfo",
            fontSize=8,
            alignment=1,  # Centralizado
            textColor=colors.grey,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SGB",
            fontSize=11,
            leading=13,
            spaceBefore=6 * mm,  # Ajustado espaçamento
            textColor=colors.HexColor("#004080"),
            fontName="Helvetica-Bold",
        )
    )
    styles.add(
        ParagraphStyle(
            name="Posto",
            fontSize=10,
            leading=12,
            spaceBefore=1 * mm,  # Ajustado espaçamento
            textColor=colors.HexColor("#404040"),  # Cor mais escura
        )
    )
    styles.add(
        ParagraphStyle(  # Estilo para o total por seção
            name="TotalPosto",
            fontSize=8,
            alignment=2,  # Alinhado à direita
            spaceBefore=1 * mm,
            textColor=colors.dimgray,
        )
    )

    # Configura o documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm,  # Aumenta margem inferior para footer/watermark
    )

    elements = []  # Lista de elementos do ReportLab (parágrafos, tabelas, etc.)

    # --- Cabeçalho com Logos ---
    try:
        # Função auxiliar para buscar imagens estáticas
        def get_image_path(file):
            path = finders.find(f"img/{file}")
            if not path:
                # Lança um erro se não encontrar para depuração
                raise FileNotFoundError(f"Arquivo estático img/{file} não encontrado!")
            return path

        logo_path = get_image_path("logo.png")
        brasao_path = get_image_path("brasao.png")

        logo_esquerda = Image(logo_path, width=12 * mm, height=20 * mm)
        logo_direita = Image(brasao_path, width=20 * mm, height=20 * mm)

        header_text = Paragraph(
            "Relação de Prioridade de Transferência<br/>15º Grupamento de Bombeiros Militar",
            styles["Header"],
        )

        header_table = Table(
            [[logo_esquerda, header_text, logo_direita]],
            colWidths=[doc.width * 0.15, doc.width * 0.7, doc.width * 0.15],
        )  # Usa doc.width para responsividade à margem

        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),  # Alinha verticalmente
                    ("ALIGN", (0, 0), (0, 0), "CENTER"),  # Logo esquerda
                    ("ALIGN", (1, 0), (1, 0), "CENTER"),  # Texto central
                    ("ALIGN", (2, 0), (2, 0), "CENTER"),  # Logo direita
                ]
            )
        )

        elements.append(header_table)
        elements.append(Spacer(1, 8 * mm))  # Espaçamento após cabeçalho
    except Exception as e:
        # Adiciona uma mensagem de erro ao PDF se o cabeçalho falhar
        elements.append(
            Paragraph(f"Erro ao gerar cabeçalho: {str(e)}", styles["Normal"])
        )

    # --- Processamento e Agrupamento dos Dados ---
    # (O queryset 'data' já vem ordenado da view)
    organized_data = OrderedDict()
    for item in data:
        # Usa o mapa para obter nomes curtos/humanos se disponíveis
        sgb_key = item.sgb_destino  # Chave para agrupar pode ser o código
        sgb_display = POSTO_SECAO_MAP.get(
            item.sgb_destino, item.sgb_destino
        )  # Nome para exibir
        posto_key = item.posto_secao_destino
        posto_display = POSTO_SECAO_MAP.get(
            item.posto_secao_destino, item.posto_secao_destino
        )

        # Agrupa por SGB (display) e depois por Posto (display)
        if sgb_display not in organized_data:
            organized_data[sgb_display] = OrderedDict()
        if posto_display not in organized_data[sgb_display]:
            organized_data[sgb_display][posto_display] = []
        organized_data[sgb_display][posto_display].append(item)

    total_geral = 0  # Contador para o total de solicitações

    # --- Geração das Tabelas por Seção ---
    table_header = [
        "Posto/Grad",
        "RE",
        "Nome Guerra",  #'Nome de Guerra',
        "SGB Origem",
        "Posto Origem",
        "Dt. Pedido",  #'Data Pedido',
        "Dias",
    ]

    # Estilo da tabela
    table_style = TableStyle(
        [
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#E0E0E0"),
            ),  # Cor cabeçalho
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),  # Alinhamento cabeçalho
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 3 * mm),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),  # Alinhamento corpo
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7),  # Tamanho fonte corpo reduzido
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.5 * mm),  # Padding interno
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.5 * mm),
            ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
        ]
    )

    # Itera sobre os dados agrupados
    for sgb_display, postos in organized_data.items():
        sgb_elements = []  # Elementos para manter juntos (KeepTogether)
        sgb_elements.append(Paragraph(sgb_display, styles["SGB"]))

        for posto_display, items in postos.items():
            posto_elements = []
            posto_elements.append(Paragraph(posto_display, styles["Posto"]))
            posto_elements.append(Spacer(1, 1.5 * mm))

            table_data = [table_header]  # Inicia com cabeçalho
            count_posto = 0
            for item in items:
                # Acessa dados relacionados com segurança (verifica existência)
                # Usa strip_tags para remover HTML de `grad` se houver
                posto_grad = (
                    strip_tags(item.cadastro.promocoes.last().grad)
                    if hasattr(item.cadastro, "promocoes")
                    and item.cadastro.promocoes.exists()
                    else "-"
                )

                # Acessa detalhes da situação com segurança
                last_situacao = (
                    item.cadastro.detalhes_situacao.last()
                    if hasattr(item.cadastro, "detalhes_situacao")
                    and item.cadastro.detalhes_situacao.exists()
                    else None
                )

                sgb_origem = (
                    POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb)
                    if last_situacao
                    else "-"
                )
                posto_origem = (
                    POSTO_SECAO_MAP.get(
                        last_situacao.posto_secao, last_situacao.posto_secao
                    )
                    if last_situacao
                    else "-"
                )

                table_data.append(
                    [
                        posto_grad,
                        f"{item.cadastro.re}-{item.cadastro.dig}",
                        item.cadastro.nome_de_guerra,
                        sgb_origem,
                        posto_origem,
                        item.data_pedido.strftime("%d/%m/%Y"),
                        str(
                            (datetime.now().date() - item.data_pedido).days
                        ),  # Calcula dias
                    ]
                )
                count_posto += 1

            # Cria a tabela para este posto/seção
            table = Table(
                table_data,
                colWidths=[
                    25 * mm,
                    15 * mm,
                    40 * mm,
                    20 * mm,
                    55 * mm,
                    20 * mm,
                    10 * mm,
                ],  # Ajuste larguras
                style=table_style,
            )

            posto_elements.append(table)
            posto_elements.append(
                Paragraph(f"Total {posto_display}: {count_posto}", styles["TotalPosto"])
            )
            posto_elements.append(Spacer(1, 4 * mm))  # Espaço após cada posto

            # Adiciona os elementos do posto ao bloco do SGB
            sgb_elements.extend(posto_elements)
            total_geral += count_posto  # Incrementa total geral

        # Adiciona o bloco do SGB (com todos os seus postos) para manter junto na página
        elements.append(KeepTogether(sgb_elements))
        elements.append(Spacer(1, 6 * mm))  # Espaço entre blocos de SGB

    # --- Funções para Rodapé e Marca D'água ---
    # (Definidas dentro para ter acesso a 'request' e 'total_geral')

    def footer(canvas, doc):
        """Adiciona o rodapé padrão em cada página."""
        canvas.saveState()
        canvas.setFont("Helvetica", 8)

        # Obter informações do usuário logado de forma segura
        posto_grad = getattr(getattr(user, "profile", None), "posto_grad", "")
        re = getattr(getattr(user, "profile", None), "re", "")
        dig = getattr(getattr(user, "profile", None), "dig", "")
        last_name = getattr(user, "last_name", "")
        user_info = f"{posto_grad} {re}-{dig} {last_name}".strip()
        if not user_info:  # Fallback para username se profile não existir
            user_info = user.username

        footer_text = (
            f"Emitido por: {user_info} | "
            f"{datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"Página {doc.page} | "
            f"Total de Solicitações: {total_geral}"
        )
        # Desenha o texto centralizado na parte inferior
        canvas.drawCentredString(A4[0] / 2, 10 * mm, footer_text)  # Ajusta posição Y
        canvas.restoreState()

    def add_watermark(canvas, doc):
        """Adiciona a marca d'água com CPF em cada página."""
        canvas.saveState()
        canvas.setFont("Helvetica", 40)  # Tamanho da fonte da marca d'água
        canvas.setFillGray(0.9, 0.15)  # Cor cinza claro e transparência (0.15 = 15%)

        # Tenta obter o CPF do profile do usuário
        cpf = getattr(
            getattr(user, "profile", None), "cpf", ""
        )  # Obtem CPF ou string vazia
        text = cpf if cpf else "CONFIDENCIAL"  # Texto da marca d'água (fallback)

        angle = 45  # Ângulo da diagonal

        # Desenha a marca d'água repetidamente
        for x in range(
            -int(A4[1]), int(A4[0] * 1.5), 250
        ):  # Ajusta range e espaçamento X
            for y in range(
                -int(A4[0]), int(A4[1] * 1.5), 150
            ):  # Ajusta range e espaçamento Y
                canvas.saveState()
                # Rotaciona o canvas em torno do ponto (x, y)
                canvas.translate(x, y)
                canvas.rotate(angle)
                # Desenha o texto centralizado na origem (0,0) após rotação
                canvas.drawCentredString(0, 0, text)
                canvas.restoreState()
        canvas.restoreState()

    def on_page(canvas, doc):
        """Função chamada pelo ReportLab para cada página."""
        footer(canvas, doc)  # Desenha o rodapé primeiro
        add_watermark(canvas, doc)  # Desenha a marca d'água por cima

    # --- Construção do PDF ---
    # Usa a função on_page para aplicar rodapé e marca d'água em todas as páginas
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    # Retorna o buffer e o nome do arquivo sugerido
    pdf_buffer = buffer.getvalue()
    buffer.close()
    return pdf_buffer, "relatorio_prioridade_transferencia.pdf"


# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO EXCEL (Mantida como estava)
# ==============================================
def export_to_excel_rpt(request, data):
    buffer = io.BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "RPT - Prioridade Transferência"  # Nome da aba

    # Cabeçalho do arquivo Excel
    ws.append(["RELATÓRIO DE PRIORIDADE DE TRANSFERÊNCIA - 15º GBM"])
    ws.append([f"Data de emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    # Obtem nome do usuário de forma segura
    emitter = request.user.get_full_name() or request.user.username
    ws.append([f"Emitido por: {emitter}"])
    ws.append([])  # Linha em branco

    # Cabeçalho da tabela
    headers = [
        "Posto/Grad",
        "RE",
        "Nome Guerra",
        "SGB Origem",
        "Posto Origem",
        "Data Pedido",
        "Dias Espera",  # Nome da coluna ajustado
    ]
    ws.append(headers)

    # Preenche as linhas com dados
    for item in data:
        # Acessa dados relacionados com segurança
        posto_grad = (
            strip_tags(item.cadastro.promocoes.last().grad)
            if hasattr(item.cadastro, "promocoes") and item.cadastro.promocoes.exists()
            else "-"
        )
        last_situacao = (
            item.cadastro.detalhes_situacao.last()
            if hasattr(item.cadastro, "detalhes_situacao")
            and item.cadastro.detalhes_situacao.exists()
            else None
        )

        sgb_origem = (
            POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb)
            if last_situacao
            else "-"
        )
        posto_origem = (
            POSTO_SECAO_MAP.get(last_situacao.posto_secao, last_situacao.posto_secao)
            if last_situacao
            else "-"
        )

        ws.append(
            [
                posto_grad,
                f"{item.cadastro.re}-{item.cadastro.dig}",
                item.cadastro.nome_de_guerra,
                sgb_origem,
                posto_origem,
                item.data_pedido,  # Excel pode formatar a data
                (datetime.now().date() - item.data_pedido).days,  # Dias como número
            ]
        )

    # Ajusta largura das colunas (opcional)
    column_widths = {"A": 15, "B": 10, "C": 30, "D": 15, "E": 40, "F": 15, "G": 10}
    for col_letter, width in column_widths.items():
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = width

    # Formata coluna de data e dias (opcional)
    for row in ws.iter_rows(
        min_row=6, max_col=7, max_row=ws.max_row
    ):  # Começa após cabeçalhos
        if row[5].value:  # Coluna F (Data Pedido)
            row[5].number_format = "dd/mm/yyyy"
        if row[6].value is not None:  # Coluna G (Dias Espera)
            row[6].number_format = "0"  # Número inteiro

    wb.save(buffer)
    buffer_bytes = buffer.getvalue()  # Extrai os bytes
    buffer.close()

    return buffer_bytes, "relatorio_prioridade_transferencia.xlsx"


# ==============================================
# FUNÇÃO PARA EXPORTAÇÃO CSV (Mantida como estava, com ajustes)
# ==============================================
def export_to_csv_rpt(request, data):
    buffer = io.BytesIO()
    # Garante UTF-8 com BOM para compatibilidade Excel
    buffer.write(b"\xef\xbb\xbf")

    # Usa TextIOWrapper para lidar com encoding corretamente
    wrapper = io.TextIOWrapper(
        buffer,
        encoding="utf-8-sig",
        newline="",  # Importante para evitar linhas extras
        write_through=True,  # Garante escrita imediata no buffer
    )

    # Cria o escritor CSV com ; como delimitador
    writer = csv.writer(
        wrapper, delimiter=";", quoting=csv.QUOTE_MINIMAL
    )  # Usar QUOTE_MINIMAL ou ALL

    try:
        # Cabeçalho do arquivo CSV
        writer.writerow(["RELATÓRIO DE PRIORIDADE DE TRANSFERÊNCIA - 15º GBM"])
        writer.writerow(
            [f"Data de emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
        )
        emitter = request.user.get_full_name() or request.user.username
        writer.writerow([f"Emitido por: {emitter}"])
        writer.writerow([])  # Linha em branco

        # Cabeçalho da tabela
        headers = [
            "Posto/Grad",
            "RE",
            "Nome Guerra",
            "SGB Origem",
            "Posto Origem",
            "Data Pedido",
            "Dias Espera",
        ]
        writer.writerow(headers)

        # Preenche as linhas com dados
        for item in data:
            posto_grad = (
                strip_tags(item.cadastro.promocoes.last().grad)
                if hasattr(item.cadastro, "promocoes")
                and item.cadastro.promocoes.exists()
                else "-"
            )
            last_situacao = (
                item.cadastro.detalhes_situacao.last()
                if hasattr(item.cadastro, "detalhes_situacao")
                and item.cadastro.detalhes_situacao.exists()
                else None
            )

            sgb_origem = (
                POSTO_SECAO_MAP.get(last_situacao.sgb, last_situacao.sgb)
                if last_situacao
                else "-"
            )
            posto_origem = (
                POSTO_SECAO_MAP.get(
                    last_situacao.posto_secao, last_situacao.posto_secao
                )
                if last_situacao
                else "-"
            )

            writer.writerow(
                [
                    posto_grad,
                    f"{item.cadastro.re}-{item.cadastro.dig}",
                    item.cadastro.nome_de_guerra,
                    sgb_origem,
                    posto_origem,
                    item.data_pedido.strftime("%d/%m/%Y"),  # Formata data para CSV
                    (datetime.now().date() - item.data_pedido).days,  # Dias como número
                ]
            )

    finally:
        # Garante que tudo seja escrito no buffer antes de fechar
        wrapper.detach()  # Desvincula o wrapper sem fechar o buffer interno

    buffer_bytes = buffer.getvalue()  # Extrai os bytes
    buffer.close()

    return buffer_bytes, "relatorio_prioridade_transferencia.csv"
